"""Import de menú desde planilla (requisito del cliente).

Pensado para el export del sistema del local (.xls), pero también soporta
.xlsx y .csv. Detecta las columnas por encabezado:
  - nombre       ← "Descripción"
  - precio       ← "Precio 1"   (el precio base a cargar en la carta)
  - categoría    ← "Familia"
  - exclusión    ← "Stock Mínimo": las filas con valor 1 NO se cargan a la carta.
Limpia precios en formato argentino ('$14.500,00' -> 14500) y al importar
reemplaza el menú completo (el nuevo pisa al viejo).
"""
from __future__ import annotations

import csv
import io
import re

from openpyxl import load_workbook

# Palabras clave por campo, en orden de prioridad; match por substring sobre el
# encabezado en minúscula.
COLUMNAS = {
    'nombre':     ['descripción', 'descripcion', 'nombre', 'producto', 'plato', 'detalle', 'artículo', 'articulo'],
    'precio':     ['precio 1', 'precio1', 'precio', 'importe', 'valor', 'monto'],
    'categoria':  ['familia', 'categoría', 'categoria', 'rubro', 'tipo'],
    'excluir':    ['stock mínimo', 'stock minimo', 'excluir', 'no mostrar', 'oculto'],
    'disponible': ['disponible', 'activo', 'hay'],   # 'stock' a propósito NO: choca con Stock Mínimo/Trabajo
}

_TRUTHY = {'1', 'si', 'sí', 's', 'yes', 'y', 'true', 'verdadero', 'hay', 'disponible', 'activo', 'ok'}


def parse_menu(filename: str, data: bytes) -> tuple[list[dict], dict]:
    """Devuelve (items, meta). Lanza ValueError si no detecta nombre/precio."""
    headers, filas = _leer_filas(filename, data)
    mapping = _detectar_columnas(headers)
    if 'nombre' not in mapping or 'precio' not in mapping:
        raise ValueError("No pude detectar las columnas de nombre y precio. Revisá los encabezados.")

    items, excluidas, descartadas = [], 0, 0
    for fila in filas:
        if 'excluir' in mapping and _es_uno(fila.get(mapping['excluir'])):
            excluidas += 1
            continue
        nombre = str(fila.get(mapping['nombre']) or '').strip()
        precio = limpiar_precio(fila.get(mapping['precio']))
        if not nombre or precio is None or precio <= 0:
            descartadas += 1
            continue
        items.append({
            'nombre': nombre,
            'categoria': str(fila.get(mapping.get('categoria', '')) or '').strip() or 'General',
            'precio': precio,
            'disponible': _parse_disponible(fila.get(mapping['disponible'])) if 'disponible' in mapping else True,
        })
    return items, {'columnas': mapping, 'total': len(items),
                   'excluidas': excluidas, 'descartadas': descartadas}


def limpiar_precio(raw) -> int | None:
    """'$14.500,00' -> 14500. Devuelve pesos enteros (sin decimales) o None."""
    s = re.sub(r'[^\d.,\-]', '', str(raw or '').strip())
    if not s:
        return None
    neg = s.startswith('-')
    s = s.lstrip('-')
    if ',' in s:                       # coma = decimal, punto = miles
        s = s.replace('.', '').replace(',', '.')
    elif '.' in s:                     # ambiguo: si los grupos son de 3 dígitos, son miles
        grupos = s.split('.')
        if all(len(g) == 3 for g in grupos[1:]):
            s = ''.join(grupos)
    try:
        val = int(round(float(s)))
    except ValueError:
        return None
    return -val if neg else val


def plantilla_csv() -> str:
    return (
        "nombre,categoria,precio,disponible\n"
        "Milanesa napolitana,Platos,\"$8.500\",si\n"
        "Provoleta,Entradas,4200,si\n"
        "Flan casero,Postres,3500,no\n"
    )


def _es_uno(valor) -> bool:
    """True si la celda vale 1 (1, 1.0, '1', '1.0'). Marca de exclusión."""
    s = str(valor or '').strip()
    try:
        return float(s) == 1.0
    except ValueError:
        return s == '1'


def _parse_disponible(raw) -> bool:
    s = str(raw or '').strip().lower()
    return True if not s else s in _TRUTHY


def _leer_filas(filename: str, data: bytes) -> tuple[list[str], list[dict]]:
    name = (filename or '').lower()
    if name.endswith('.csv'):
        texto = data.decode('utf-8-sig', errors='replace')
        matriz = list(csv.reader(io.StringIO(texto)))
    elif name.endswith(('.xlsx', '.xlsm')):
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        matriz = [list(fila) for fila in wb.active.iter_rows(values_only=True)]
    elif name.endswith('.xls'):
        import xlrd
        hoja = xlrd.open_workbook(file_contents=data).sheet_by_index(0)
        matriz = [hoja.row_values(r) for r in range(hoja.nrows)]
    else:
        raise ValueError("Formato no soportado. Subí un .xls, .xlsx o .csv.")

    matriz = [f for f in matriz if any(c not in (None, '') for c in f)]
    if not matriz:
        return [], []

    headers = [str(c).strip().lower() if c is not None else '' for c in matriz[0]]
    filas = [
        {h: (fila[i] if i < len(fila) else None) for i, h in enumerate(headers)}
        for fila in matriz[1:]
    ]
    return headers, filas


def _detectar_columnas(headers: list[str]) -> dict:
    """Por cada campo, elige el primer encabezado que matchea la palabra clave
    más específica disponible (prioridad por el orden de COLUMNAS)."""
    mapping = {}
    for campo, claves in COLUMNAS.items():
        for clave in claves:
            hit = next((h for h in headers if h and clave in h), None)
            if hit:
                mapping[campo] = hit
                break
    return mapping
